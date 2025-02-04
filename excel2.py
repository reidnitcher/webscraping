from telnetlib import PRAGMA_HEARTBEAT
import openpyxl as xl

wb = xl.load_workbook('2_example.xlsx')

sn = wb.sheetnames

print(sn)

sheet1 = wb['Sheet1']
cellA1 = sheet1['A1']

print(sheet1)
print(cellA1.value)

print(cellA1.row)
print(cellA1.column)
print(cellA1.coordinate)

print(sheet1.cell(1,2).value)

for x in range (1,8):
    print(sheet1.cell(x,2).value)

print(sheet1.max_row)
print(sheet1.max_column)

print(xl.utils.get_column_letter(1))
print(xl.utils.get_column_letter(900))


print(xl.utils.column_index_from_string('AHP'))

for currentrow in sheet1['A1':'C3']:
    #print(currentrow)
    for currentcell in currentrow  :
        print(currentcell.coordinate, currentcell.value)
        print('---END OF COLUMNS---')

        print()
        print('---END OF ROW---')
        print()

for currentrow in sheet1.iter_rows(min_row = 1, max_row=sheet1.max_row, max_col=sheet1.max_col):
    print(currentrow[0].value)
    print(currentrow[1].value)
    print(currentrow[2].value)


